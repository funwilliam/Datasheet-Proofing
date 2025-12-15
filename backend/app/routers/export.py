# backend/app/routers/export.py
from __future__ import annotations

from fastapi import APIRouter, Depends, Response, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Dict, Any, Optional, Iterable, Literal
from datetime import datetime, timezone, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import or_, case

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from urllib.parse import quote

import io
import csv
import json

from ..db import get_db
from ..models import ModelItem

router = APIRouter(prefix="/api/export", tags=["export"])


# ──────────────────────────────────────────────────────────────────────────────
# 檔名 / Header helpers
# ──────────────────────────────────────────────────────────────────────────────

_TZ_TAIPEI = timezone(timedelta(hours=8))


def _ts_taipei() -> str:
    """UTC+8 的時間戳，方便使用者辨識不同次輸出。"""
    return datetime.now(_TZ_TAIPEI).strftime("%Y%m%d_%H%M%S")


def _content_disposition(filename_utf8: str, ascii_fallback: str) -> str:
    """
    支援中文檔名：
      - filename=        (ASCII fallback)
      - filename*=UTF-8'' (RFC 5987)
    """
    return f'attachment; filename="{ascii_fallback}"; filename*=UTF-8\'\'{quote(filename_utf8)}'


# ──────────────────────────────────────────────────────────────────────────────
# 共用工具
# ──────────────────────────────────────────────────────────────────────────────

def _dt_to_iso_z(dt: Optional[datetime]) -> Optional[str]:
    """將 datetime 轉成 ISO8601（UTC、結尾 Z）。None 則回 None。"""
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt.replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _excel_safe_cell(v: Any) -> str:
    """
    Excel formula injection 防護：
    只要以 = + - @ 開頭，前面補一個單引號。
    """
    s = "" if v is None else str(v)
    if not s:
        return ""
    if s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _serialize_model_to_json(m: ModelItem) -> Dict[str, Any]:
    """JSON 匯出用：與 /api/models/{model_number} 類似，但偏批次匯出格式。"""
    apps = [t.app_tag for t in (m.applications or [])]

    files_out: List[Dict[str, Any]] = []
    try:
        files_sorted = sorted(list(m.files or []), key=lambda fa: (
            fa.created_at or 0), reverse=True)
    except Exception:
        files_sorted = list(m.files or [])

    for fa in files_sorted:
        files_out.append({"file_hash": fa.file_hash, "filename": fa.filename})

    return {
        "model_number": m.model_number,
        "input_voltage_range": m.input_voltage_range,
        "output_voltage": m.output_voltage,
        "output_power": m.output_power,
        "package": m.package,
        "isolation": m.isolation,
        "insulation": m.insulation,
        "dimension": m.dimension,
        "applications": apps,
        "verify_status": m.verify_status,
        "reviewer": m.reviewer,
        "reviewed_at": _dt_to_iso_z(m.reviewed_at),
        "files": files_out,
    }


def _serialize_model_to_csv_row(m: ModelItem) -> Dict[str, str]:
    """
    CSV/XLSX 規格匯出用：
    - file_hashes / filenames 都輸出成「JSON array string」
      例：["hash1","hash2"] / ["a.pdf","b.pdf"]
    - 兩欄以 index 對齊，方便未來還原回 JSON
    """
    apps = [t.app_tag for t in (m.applications or [])]

    try:
        files_sorted = sorted(list(m.files or []), key=lambda fa: (
            fa.created_at or 0), reverse=True)
    except Exception:
        files_sorted = list(m.files or [])

    file_hashes = [(fa.file_hash or "") for fa in files_sorted]
    filenames = [(fa.filename or "") for fa in files_sorted]

    return {
        "model_number": m.model_number or "",
        "input_voltage_range": m.input_voltage_range or "",
        "output_voltage": m.output_voltage or "",
        "output_power": m.output_power or "",
        "package": m.package or "",
        "isolation": m.isolation or "",
        "insulation": m.insulation or "",
        "dimension": m.dimension or "",
        "applications": "; ".join(apps),
        "verify_status": (m.verify_status or ""),
        "reviewer": (m.reviewer or ""),
        "reviewed_at": (_dt_to_iso_z(m.reviewed_at) or ""),
        "file_hashes": json.dumps(file_hashes, ensure_ascii=False),
        "filenames": json.dumps(filenames, ensure_ascii=False),
    }


def _chunked_in_filter(q, column, values: List[str], chunk_size: int = 900):
    """避開 SQLite 999 參數上限：IN 拆塊後用 OR 串接。"""
    cleaned = [v for v in (values or []) if isinstance(v, str) and v.strip()]
    if not cleaned:
        return q.filter(False)
    ors = []
    for i in range(0, len(cleaned), chunk_size):
        chunk = cleaned[i: i + chunk_size]
        ors.append(column.in_(chunk))
    return q.filter(or_(*ors))


def _unique_in_order(seq: Iterable[str]) -> List[str]:
    """去重但保留第一次出現的順序；自動去掉空字串/None。"""
    seen = set()
    out: List[str] = []
    for x in seq or []:
        if not isinstance(x, str):
            continue
        s = x.strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


_SPEC_FIELDNAMES = [
    "model_number",
    "input_voltage_range",
    "output_voltage",
    "output_power",
    "package",
    "isolation",
    "insulation",
    "dimension",
    "applications",
    "verify_status",
    "reviewer",
    "reviewed_at",
    "file_hashes",
    "filenames",
]


def _csv_stream(rows: Iterable[ModelItem]) -> Iterable[bytes]:
    """串流產出 CSV bytes。"""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_SPEC_FIELDNAMES)
    writer.writeheader()
    yield buf.getvalue().encode("utf-8-sig")
    buf.seek(0)
    buf.truncate(0)

    for m in rows:
        writer.writerow(_serialize_model_to_csv_row(m))
        yield buf.getvalue().encode("utf-8")
        buf.seek(0)
        buf.truncate(0)


def _json_bytes(data: Any) -> bytes:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def _xlsx_bytes_for_specs(rows: List[ModelItem]) -> io.BytesIO:
    """
    規格匯出 XLSX：
    - 欄位與 CSV 同一套（_SPEC_FIELDNAMES）
    - file_hashes / filenames 仍為 JSON array string（可逆）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "model_specs"

    ws.append(_SPEC_FIELDNAMES)
    for m in rows:
        r = _serialize_model_to_csv_row(m)
        ws.append([_excel_safe_cell(r.get(k, "")) for k in _SPEC_FIELDNAMES])

    ws.freeze_panes = "A2"

    # 實用的欄寬（保守，不追求完美）
    widths = [
        24, 18, 14, 14, 14, 12, 12, 16, 22, 14, 16, 22, 40, 40
    ]
    for i, w_ in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w_

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ──────────────────────────────────────────────────────────────────────────────
# 整庫匯出
# ──────────────────────────────────────────────────────────────────────────────

@router.get("")
def export_data(
    status: Optional[str] = None,
    fmt: str = "json",  # 'json' | 'csv' | 'xlsx'
    db: Session = Depends(get_db),
):
    q = db.query(ModelItem)
    if status:
        q = q.filter(ModelItem.verify_status == status)

    q = q.order_by(ModelItem.model_number.asc())
    rows: List[ModelItem] = q.all()

    ts = _ts_taipei()
    label = "型號資料匯出"
    if status == "verified":
        label += "_已驗證"
    elif status == "unverified":
        label += "_未驗證"

    if fmt.lower() == "json":
        data = [_serialize_model_to_json(m) for m in rows]
        filename_utf8 = f"{label}_{ts}.json"
        headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"models_export_{ts}.json"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(
            io.BytesIO(_json_bytes(data)),
            media_type="application/json; charset=utf-8",
            headers=headers,
        )

    if fmt.lower() == "csv":
        filename_utf8 = f"{label}_{ts}.csv"
        headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"models_export_{ts}.csv"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(_csv_stream(rows), media_type="text/csv; charset=utf-8", headers=headers)

    if fmt.lower() == "xlsx":
        out = _xlsx_bytes_for_specs(rows)
        filename_utf8 = f"{label}_{ts}.xlsx"
        headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"models_export_{ts}.xlsx"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    raise HTTPException(
        status_code=400, detail="unsupported fmt (use 'json' / 'csv' / 'xlsx')")


# ──────────────────────────────────────────────────────────────────────────────
# 指定型號清單匯出（規格資料 JSON/CSV/XLSX）
# ──────────────────────────────────────────────────────────────────────────────

class ExportByModelsIn(BaseModel):
    model_numbers: List[str] = Field(..., description="要匯出的型號清單")
    status: Optional[str] = Field(
        None, description="過濾 verify_status（例：verified/unverified）；None 表示不過濾")
    fmt: Literal["json", "csv", "xlsx"] = Field("json", description="輸出格式")
    preserve_order: bool = Field(False, description="是否依照輸入清單的順序輸出")


@router.post("/by-models")
def export_by_models(payload: ExportByModelsIn, db: Session = Depends(get_db)):
    model_numbers = _unique_in_order(payload.model_numbers)

    ts = _ts_taipei()
    label = "型號規格匯出"
    if payload.status == "verified":
        label += "_已驗證"
    elif payload.status == "unverified":
        label += "_未驗證"

    # 空清單：仍回「可下載」檔（header-only），UX 比較一致
    if not model_numbers:
        if payload.fmt.lower() == "json":
            filename_utf8 = f"{label}_{ts}.json"
            headers = {
                "Content-Disposition": _content_disposition(filename_utf8, f"models_export_selected_{ts}.json"),
                "Cache-Control": "no-store",
            }
            return StreamingResponse(
                io.BytesIO(_json_bytes([])),
                media_type="application/json; charset=utf-8",
                headers=headers,
            )

        if payload.fmt.lower() == "csv":
            filename_utf8 = f"{label}_{ts}.csv"
            headers = {
                "Content-Disposition": _content_disposition(filename_utf8, f"models_export_selected_{ts}.csv"),
                "Cache-Control": "no-store",
            }
            # 只有 BOM + header

            def _empty_csv():
                buf = io.StringIO()
                w = csv.DictWriter(buf, fieldnames=_SPEC_FIELDNAMES)
                w.writeheader()
                yield buf.getvalue().encode("utf-8-sig")

            return StreamingResponse(_empty_csv(), media_type="text/csv; charset=utf-8", headers=headers)

        if payload.fmt.lower() == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "model_specs"
            ws.append(_SPEC_FIELDNAMES)
            ws.freeze_panes = "A2"

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            filename_utf8 = f"{label}_{ts}.xlsx"
            headers = {
                "Content-Disposition": _content_disposition(filename_utf8, f"models_export_selected_{ts}.xlsx"),
                "Cache-Control": "no-store",
            }
            return StreamingResponse(
                out,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers,
            )

        raise HTTPException(
            status_code=400, detail="unsupported fmt (use 'json' / 'csv' / 'xlsx')")

    # 有資料：查 DB
    q = db.query(ModelItem)
    q = _chunked_in_filter(q, ModelItem.model_number, model_numbers)

    if payload.status:
        q = q.filter(ModelItem.verify_status == payload.status)

    if payload.preserve_order:
        order_case = case(
            *[(ModelItem.model_number == mn, idx)
              for idx, mn in enumerate(model_numbers)],
            else_=len(model_numbers),
        )
        q = q.order_by(order_case)
    else:
        q = q.order_by(ModelItem.model_number.asc())

    rows: List[ModelItem] = q.all()

    # JSON 若 preserve_order，再保險一次
    if payload.preserve_order and payload.fmt.lower() == "json":
        idx = {mn: i for i, mn in enumerate(model_numbers)}
        rows.sort(key=lambda m: idx.get(m.model_number, len(idx) + 1))

    if payload.fmt.lower() == "json":
        data = [_serialize_model_to_json(m) for m in rows]
        filename_utf8 = f"{label}_{ts}.json"
        headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"models_export_selected_{ts}.json"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(
            io.BytesIO(_json_bytes(data)),
            media_type="application/json; charset=utf-8",
            headers=headers,
        )

    if payload.fmt.lower() == "csv":
        filename_utf8 = f"{label}_{ts}.csv"
        headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"models_export_selected_{ts}.csv"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(_csv_stream(rows), media_type="text/csv; charset=utf-8", headers=headers)

    if payload.fmt.lower() == "xlsx":
        # preserve_order 若為 True：query 已排序；這裡不再重排
        out = _xlsx_bytes_for_specs(rows)
        filename_utf8 = f"{label}_{ts}.xlsx"
        headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"models_export_selected_{ts}.xlsx"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    raise HTTPException(
        status_code=400, detail="unsupported fmt (use 'json' / 'csv' / 'xlsx')")


# ──────────────────────────────────────────────────────────────────────────────
# UX 友善 Model Settings 模板匯出
# 欄位：型號 / 系列 / 製造商 / Mouser / DigiKey / Arrow / Future / RS產品網址
# 一列一型號
# ──────────────────────────────────────────────────────────────────────────────

class ExportModelSettingsIn(BaseModel):
    model_numbers: List[str] = Field(..., description="要匯出的型號清單")
    fmt: Literal["csv", "xlsx"] = Field("xlsx", description="輸出格式")
    preserve_order: bool = Field(True, description="是否依照輸入清單的順序輸出")


@router.post("/model-settings")
def export_model_settings(payload: ExportModelSettingsIn, db: Session = Depends(get_db)):
    model_numbers = _unique_in_order(payload.model_numbers)

    headers = ["型號", "系列", "製造商", "Mouser",
               "DigiKey", "Arrow", "Future", "RS產品網址"]

    rows: List[List[str]] = []
    for mn in model_numbers:
        rows.append([mn, "", "", "", "", "", "", ""])

    ts = _ts_taipei()
    base_label = "型號設定模板"

    if payload.fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow([_excel_safe_cell(c) for c in r])

        data = buf.getvalue().encode("utf-8-sig")
        filename_utf8 = f"{base_label}_{ts}.csv"
        resp_headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"model_settings_{ts}.csv"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(io.BytesIO(data), media_type="text/csv; charset=utf-8", headers=resp_headers)

    if payload.fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "model_settings"

        ws.append(headers)
        for r in rows:
            ws.append([_excel_safe_cell(c) for c in r])

        ws.freeze_panes = "A2"

        widths = [22, 14, 18, 10, 10, 10, 10, 40]
        for i, w_ in enumerate(widths, start=1):
            ws.column_dimensions[chr(ord("A") + i - 1)].width = w_

        last_row = max(2, len(rows) + 1)
        dv_bool = DataValidation(
            type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        ws.add_data_validation(dv_bool)
        dv_bool.add(f"D2:G{last_row}")  # Mouser..Future

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename_utf8 = f"{base_label}_{ts}.xlsx"
        resp_headers = {
            "Content-Disposition": _content_disposition(filename_utf8, f"model_settings_{ts}.xlsx"),
            "Cache-Control": "no-store",
        }
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=resp_headers,
        )

    raise HTTPException(status_code=400, detail="unsupported fmt (use 'csv' or 'xlsx')")
